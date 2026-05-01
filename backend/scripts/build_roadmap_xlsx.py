"""
Generates docs/30_Day_Development_Roadmap.xlsx — multi-sheet ownership plan.
Run: python scripts/build_roadmap_xlsx.py
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import date, timedelta

OUT = "docs/30_Day_Development_Roadmap.xlsx"

# ---------- styling helpers ----------
THIN = Side(border_style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

PHASE_COLORS = {
    "Phase 1 – Stabilize": "FCE4D6",
    "Phase 2 – Harden":    "FFF2CC",
    "Phase 3 – Optimize":  "DDEBF7",
    "Phase 4 – Advance":   "E2EFDA",
}
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT  = Font(bold=True, size=16, color="1F4E78")
SUB_FONT    = Font(bold=True, size=12, color="1F4E78")

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = BORDER

def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

def working_day(start: date, n: int) -> date:
    """Return the nth working day starting from `start` (n=1 means start)."""
    d = start
    count = 0
    while True:
        if d.weekday() < 5:
            count += 1
            if count == n:
                return d
        d += timedelta(days=1)

# ---------- workbook ----------
wb = Workbook()

# ============================================================
# Sheet 1: Cover / Executive Summary
# ============================================================
cover = wb.active
cover.title = "Executive Summary"

cover["A1"] = "CCP-Backend — 30-Day Development Roadmap"
cover["A1"].font = TITLE_FONT
cover.merge_cells("A1:F1")

cover["A2"] = "Owner: Incoming Lead Developer  |  Start: 2026-05-01  |  End: 2026-06-11 (30 working days)"
cover["A2"].font = Font(italic=True, color="595959")
cover.merge_cells("A2:F2")

cover["A4"] = "Mission"
cover["A4"].font = SUB_FONT
cover["A5"] = (
    "Take ownership of a working Flask payments backend, eliminate critical money-flow risks, "
    "establish production-grade observability and CI, and ship one new feature end-to-end. "
    "The codebase is healthy — this plan stabilizes, hardens, optimizes, then advances."
)
cover["A5"].alignment = Alignment(wrap_text=True, vertical="top")
cover.merge_cells("A5:F8")

cover["A10"] = "Phases"
cover["A10"].font = SUB_FONT

phases = [
    ("Phase 1 – Stabilize",  "Days 1–7",   "Run locally, fix webhook idempotency + balance race, wire metrics & Sentry."),
    ("Phase 2 – Harden",     "Days 8–14",  "CI pipeline, structured logging, rate limiting, request IDs, security pass."),
    ("Phase 3 – Optimize",   "Days 15–21", "Query audit, Redis caching, async webhooks, load test, APM tracing."),
    ("Phase 4 – Advance",    "Days 22–30", "Reconciliation job, fraud rules, MFA, audit log, docs, demo."),
]
cover.append([])  # row 11
style_header(cover, 11, 3)
cover.cell(row=11, column=1, value="Phase")
cover.cell(row=11, column=2, value="Window")
cover.cell(row=11, column=3, value="Focus")
for i, (name, window, focus) in enumerate(phases, start=12):
    cover.cell(row=i, column=1, value=name).fill = PatternFill("solid", fgColor=PHASE_COLORS[name])
    cover.cell(row=i, column=2, value=window)
    cover.cell(row=i, column=3, value=focus).alignment = Alignment(wrap_text=True, vertical="top")
    for c in range(1, 4):
        cover.cell(row=i, column=c).border = BORDER

cover["A18"] = "Success Criteria (end of Day 30)"
cover["A18"].font = SUB_FONT
crit = [
    "✓ Zero known race conditions in money flow (idempotency + atomic balance updates verified by concurrency tests)",
    "✓ Sentry capturing prod errors; one consolidated metrics module exporting to /metrics; structured JSON logs in prod",
    "✓ CI pipeline green on every PR (lint, type, test ≥ 80% coverage on modules/wallet and modules/community)",
    "✓ Load test baseline established: documented RPS at p95 < 500ms for /wallet and /community endpoints",
    "✓ Reconciliation job running daily; audit log captures every wallet credit/debit",
    "✓ One new feature shipped end-to-end through the modules/* stack (resource → schema → service → repo → migration → test)",
    "✓ 60-day plan drafted from learnings",
]
for i, line in enumerate(crit, start=19):
    cover.cell(row=i, column=1, value=line)
    cover.merge_cells(start_row=i, start_column=1, end_row=i, end_column=6)

cover["A27"] = "Legend"
cover["A27"].font = SUB_FONT
cover["A28"] = "Priority: 🔴 Critical  🟠 High  🟡 Medium  🟢 Low"
cover["A29"] = "Effort: hours of focused work (excludes meetings, code review)"
cover["A30"] = "Dependencies: Day numbers that must complete first"

autosize(cover, [22, 14, 70, 14, 14, 14])
cover.row_dimensions[5].height = 60
for r in (12, 13, 14, 15):
    cover.row_dimensions[r].height = 32

# ============================================================
# Sheet 2: 30-Day Roadmap (the main sheet)
# ============================================================
plan = wb.create_sheet("30-Day Roadmap")

headers = [
    "Day", "Date", "Phase", "Category", "Task",
    "Subtasks / Implementation Notes", "Files / Areas",
    "Priority", "Effort (h)", "Depends On",
    "Acceptance Criteria", "Risk", "Status",
]

# Day-by-day. Each row: tuple matching headers (Day, Date filled later, ...)
START = date(2026, 5, 1)

ROWS = [
    # ---------------- Phase 1: Stabilize ----------------
    (1, "Phase 1 – Stabilize", "Onboarding",
     "Local environment + repo walkthrough",
     "Clone, copy .env.example → .env, fill secrets (Bell MFB sandbox, SafeHaven sandbox, Paystack test keys, Redis URL). "
     "Run: docker compose up. Verify /health returns 200. Run pytest, capture baseline coverage with --cov=modules. "
     "Read CLAUDE.md, README.md, and docs/MATURITY_ASSESSMENT.md cover-to-cover.",
     "docker-compose.yml, .env.example, README.md",
     "🟠 High", 6, "—",
     "App boots locally; pytest runs; coverage report saved as baseline_coverage.html.",
     "Low", "Not Started"),

    (2, "Phase 1 – Stabilize", "Discovery",
     "Money-path read-through + threat model",
     "Read modules/wallet/resources/webhook_resource.py end-to-end. Trace a Bell MFB credit from webhook → service → "
     "repository → DB. Same for SafeHaven path. Sketch sequence diagram. List every db.session.commit() in money paths "
     "and ask: what happens if this is called twice concurrently?",
     "modules/wallet/**, modules/community/services/payment_service*.py",
     "🟠 High", 7, "1",
     "Sequence diagram + 1-page threat model committed to docs/money_flow_threat_model.md.",
     "Low", "Not Started"),

    (3, "Phase 1 – Stabilize", "Bug Fix",
     "Webhook idempotency: enforce DB-level uniqueness",
     "Generate Alembic migration adding UNIQUE constraint on wallet_transactions.bell_mfb_reference (and the SafeHaven "
     "equivalent). Wrap the insert in try/except IntegrityError and treat as duplicate (return 200). Keep the existing "
     "pre-check as fast path. Add migration to flask_migrations/versions/.",
     "modules/wallet/resources/webhook_resource.py:142-154, modules/wallet/repositories/, flask_migrations/versions/",
     "🔴 Critical", 6, "2",
     "New migration applies cleanly. Concurrent insert of same reference results in exactly one row.",
     "Medium", "Not Started"),

    (4, "Phase 1 – Stabilize", "Bug Fix",
     "Atomic wallet balance updates",
     "Audit wallet_repo.update_balance(). If it does Python read-modify-write, replace with SQL increment: "
     "UPDATE wallets SET balance = balance + :delta WHERE id = :id RETURNING balance. Same audit on community wallet "
     "balance updates. Use SELECT … FOR UPDATE only if a true read-then-decision is needed (e.g., debit guard).",
     "modules/wallet/repositories/wallet_repository.py, modules/wallet/utils/transaction_manager.py",
     "🔴 Critical", 7, "3",
     "Two concurrent credits of ₦100 each to the same wallet result in +₦200 — verified by integration test.",
     "High", "Not Started"),

    (5, "Phase 1 – Stabilize", "Observability",
     "Consolidate metrics + wire Sentry",
     "Delete metrics.py and metrics_manager.py; keep metrics_registry.py (or pick one). Initialize in create_app(). "
     "Expose /metrics endpoint behind admin auth. Add sentry-sdk[flask] to requirements; init in app.py with "
     "SENTRY_DSN env var, environment tag, traces_sample_rate=0.1. Add /sentry-debug route gated on DEBUG.",
     "metrics*.py, app.py, requirements.txt, requirements-prod.txt",
     "🔴 Critical", 5, "1",
     "/metrics returns Prometheus output with wallet_* counters incrementing. Test exception surfaces in Sentry.",
     "Low", "Not Started"),

    (6, "Phase 1 – Stabilize", "Testing",
     "Concurrency tests for money paths",
     "Add tests/integration/ with a real Postgres via testcontainers-python (or pytest-postgresql). Tests: "
     "(a) 50 concurrent identical webhook calls → 1 transaction row; (b) 20 concurrent ₦100 credits to same wallet → "
     "balance == ₦2000; (c) debit attempt that would overdraw → rejected, balance unchanged.",
     "tests/integration/test_wallet_concurrency.py, conftest.py, pytest.ini, requirements.txt",
     "🟠 High", 7, "3,4",
     "Three integration tests pass repeatedly (run 10x) without flakes.",
     "Medium", "Not Started"),

    (7, "Phase 1 – Stabilize", "Review",
     "Week 1 retrospective + PR merge train",
     "Open PRs for Days 3–6 (split by concern). Self-review with /review skill. Merge to main with semver tag v2.0.0-rc1. "
     "Write Week 1 retro: what surprised you, what's still scary. Update docs/MATURITY_ASSESSMENT.md.",
     "PRs, docs/MATURITY_ASSESSMENT.md",
     "🟡 Medium", 4, "3,4,5,6",
     "All Phase 1 PRs merged. Tag pushed. Retro doc committed.",
     "Low", "Not Started"),

    # ---------------- Phase 2: Harden ----------------
    (8, "Phase 2 – Harden", "Cleanup",
     "Replace print() with logger; tidy migration scripts",
     "Sweep ~70 print() calls in modules/ to logger.info/warning. Hot files: modules/verification/providers/"
     "paystack_provider.py:300-301, firebase/service.py:17,27,29. Move clear_migrations_db.py + fix_migrations.py "
     "into scripts/maintenance/ and add a README warning.",
     "modules/**/*.py, firebase/service.py, scripts/maintenance/",
     "🟡 Medium", 5, "7",
     "grep -RIn 'print(' modules/ returns only legitimate uses. Maintenance scripts moved.",
     "Low", "Not Started"),

    (9, "Phase 2 – Harden", "Infra",
     "Verify Celery wiring + worker boot",
     "Confirm celery_app exists and is imported by celery-worker container. Add a smoke task (celery_app.send_task) "
     "and a fixture in tests. If broken, build modules/tasks/celery_app.py and wire into docker-compose celery-worker. "
     "Document in README how to enqueue a job locally.",
     "modules/tasks/, docker-compose.yml, README.md",
     "🟠 High", 6, "7",
     "Local worker picks up a test task and logs completion. CI runs the same smoke test.",
     "Medium", "Not Started"),

    (10, "Phase 2 – Harden", "Infra",
     "CI pipeline (GitHub Actions)",
     "Create .github/workflows/ci.yml: matrix on Python 3.11. Steps: ruff check, mypy (best-effort), pytest with "
     "Postgres + Redis services, coverage gate ≥ 80% for modules/wallet and modules/community. PR check required.",
     ".github/workflows/ci.yml, pyproject.toml or ruff.toml",
     "🔴 Critical", 6, "6",
     "PR opened against main runs full CI; failing test blocks merge.",
     "Low", "Not Started"),

    (11, "Phase 2 – Harden", "Observability",
     "Structured JSON logging + correlation IDs",
     "Add python-json-logger. Configure root logger to emit JSON in prod (FLASK_ENV=production), text in dev. "
     "Add Flask before_request hook to attach X-Request-ID (generate if absent) into a contextvar; logger filter "
     "injects it into every log line.",
     "app.py, modules/core/logging.py (new), config.py",
     "🟠 High", 5, "5,10",
     "curl with -H 'X-Request-ID: abc' produces JSON logs all stamped request_id=abc.",
     "Low", "Not Started"),

    (12, "Phase 2 – Harden", "Security",
     "Rate limiting on auth + money endpoints",
     "Add Flask-Limiter with Redis backend. Defaults: 60/min global. Tighter on /auth/login (10/min/IP), /auth/otp "
     "(5/min/email), /wallet/transfer (20/min/user). Custom key function uses user_id when authenticated.",
     "modules/auth_v2/resources/, modules/wallet/resources/, app.py, requirements.txt",
     "🟠 High", 5, "10",
     "Hammering /auth/login from one IP returns 429 after 10 reqs; /metrics counter shows blocks.",
     "Medium", "Not Started"),

    (13, "Phase 2 – Harden", "Security",
     "Dependency + secret scan + CSP headers",
     "Add pip-audit and trufflehog to CI. Fix HIGH findings. Add Talisman for security headers (HSTS, no-sniff, "
     "frame-deny, referrer-policy). Audit CORS allowlist in config.py.",
     "config.py, app.py, .github/workflows/ci.yml, requirements.txt",
     "🟠 High", 5, "10",
     "pip-audit clean of HIGH; trufflehog clean; security headers verified via curl -I.",
     "Medium", "Not Started"),

    (14, "Phase 2 – Harden", "Review",
     "Week 2 retro + threat model update",
     "Merge train for Phase 2 PRs. Re-run threat model from Day 2 — what's now mitigated? Update "
     "docs/MATURITY_ASSESSMENT.md with new posture. Tag v2.0.0-rc2.",
     "docs/, PRs, git tag",
     "🟡 Medium", 4, "8,9,10,11,12,13",
     "Retro doc updated. Tag pushed. Phase 2 board cleared.",
     "Low", "Not Started"),

    # ---------------- Phase 3: Optimize ----------------
    (15, "Phase 3 – Optimize", "Performance",
     "Query audit (N+1 hunt + missing indexes)",
     "Enable SQLAlchemy echo in a profiling test run. Run k6 against /community/feed, /wallet/transactions, "
     "/auth/me — capture query counts. Use EXPLAIN ANALYZE on slow ones. Output: docs/query_audit.md with "
     "per-endpoint findings + proposed indexes.",
     "modules/**/repositories/, docs/query_audit.md",
     "🟠 High", 7, "10",
     "Audit doc lists ≥ 5 N+1s or missing indexes with file:line references.",
     "Low", "Not Started"),

    (16, "Phase 3 – Optimize", "Performance",
     "Add indexes + fix top 3 N+1 queries",
     "Migration adding indexes from Day 15 audit. Refactor top 3 N+1s using joinedload / selectinload. Re-run "
     "k6 profile and compare query counts before/after.",
     "modules/**/repositories/, flask_migrations/versions/",
     "🟠 High", 7, "15",
     "Each fixed endpoint shows ≥ 50% query-count reduction. Indexes applied in dev + staging.",
     "Medium", "Not Started"),

    (17, "Phase 3 – Optimize", "Performance",
     "Redis caching for read-hot endpoints",
     "Add cache layer (cachetools + Redis). Wrap expensive reads: GET /community/<id>, GET /wallet/<id>/balance "
     "(with short TTL + write-through invalidation on transaction). Cache-key includes user_id.",
     "modules/core/cache.py (new), modules/community/services/, modules/wallet/services/",
     "🟡 Medium", 6, "16",
     "Cache hit ratio > 70% on /community/<id> under load test. Invalidation verified after a write.",
     "Medium", "Not Started"),

    (18, "Phase 3 – Optimize", "Reliability",
     "Async webhook processing via Celery",
     "Webhook endpoint accepts → validates signature → enqueues processing task → returns 200 immediately. "
     "Celery task does the DB work with the same idempotency guarantees. Add DLQ for failures.",
     "modules/wallet/resources/webhook_resource.py, modules/tasks/wallet_tasks.py (new)",
     "🟠 High", 8, "9,17",
     "Webhook p95 < 100ms. Task retries on transient DB errors. DLQ surfaces poison messages.",
     "High", "Not Started"),

    (19, "Phase 3 – Optimize", "Performance",
     "DB connection pooling + tuning",
     "Tune SQLAlchemy pool_size, max_overflow, pool_pre_ping, pool_recycle for Neon. Match gunicorn workers × "
     "threads to pool. Document the math in docs/db_pool_tuning.md.",
     "config.py, app.py, Dockerfile (gunicorn args)",
     "🟡 Medium", 4, "16",
     "No 'connection invalidated' errors under sustained load. Doc explains the chosen numbers.",
     "Medium", "Not Started"),

    (20, "Phase 3 – Optimize", "Performance",
     "Load test baseline (Locust)",
     "Build locustfile.py with realistic flows: login → wallet view → transfer; community feed scroll; webhook "
     "burst. Run against staging at 50/100/200 RPS. Capture p50/p95/p99 latency + error rate. Output: "
     "docs/load_test_baseline.md.",
     "tests/load/locustfile.py, docs/load_test_baseline.md",
     "🟠 High", 6, "18,19",
     "Baseline doc with charts; team agrees on capacity ceiling for current infra.",
     "Medium", "Not Started"),

    (21, "Phase 3 – Optimize", "Observability",
     "OpenTelemetry tracing → Sentry/Honeycomb",
     "Add opentelemetry-instrumentation-flask + sqlalchemy + redis + requests. Export to OTLP. Trace a request "
     "end-to-end through webhook → Celery task → DB. Verify spans in backend.",
     "app.py, requirements.txt, modules/core/tracing.py (new)",
     "🟡 Medium", 6, "11,18",
     "Single trace ID propagates from HTTP entry through Celery task and shows DB spans.",
     "Low", "Not Started"),

    # ---------------- Phase 4: Advance ----------------
    (22, "Phase 4 – Advance", "Feature",
     "Reconciliation job — design + schema",
     "Design daily recon: pull provider transactions (Bell MFB + SafeHaven) for last 24h, diff against our DB, "
     "produce a discrepancy report. Add reconciliation_runs and reconciliation_discrepancies tables.",
     "modules/wallet/services/reconciliation_service.py (new), flask_migrations/versions/",
     "🟠 High", 6, "18",
     "Schema migration applied. Service skeleton with tests for diff logic.",
     "Medium", "Not Started"),

    (23, "Phase 4 – Advance", "Feature",
     "Reconciliation job — implementation + Celery beat",
     "Implement provider clients' list endpoints, finish diff logic, schedule via Celery beat (02:00 UTC daily). "
     "On discrepancy: alert via Sentry + email to ops. Admin endpoint to view recent runs.",
     "modules/wallet/services/reconciliation_service.py, modules/tasks/, modules/admin/resources/",
     "🟠 High", 8, "22",
     "First scheduled run completes in staging; mock discrepancy fires alert.",
     "High", "Not Started"),

    (24, "Phase 4 – Advance", "Feature",
     "Fraud rules engine (v1, deterministic)",
     "Pluggable rule chain on /wallet/transfer: velocity (>5 transfers in 1min → review), amount (>₦500k → "
     "manual review for unverified users), new device (transfer within 1h of new device login → step-up auth). "
     "Decisions logged for audit.",
     "modules/wallet/services/fraud_service.py (new), modules/wallet/models/",
     "🟠 High", 7, "23",
     "Each rule has a test. Triggered transfers create a review record visible to admins.",
     "High", "Not Started"),

    (25, "Phase 4 – Advance", "Security",
     "TOTP-based MFA opt-in",
     "Add pyotp. Endpoints: /auth/mfa/enroll (generates secret + QR), /auth/mfa/verify, /auth/mfa/disable. "
     "Login flow checks MFA enabled flag and prompts for code. Recovery codes (10 single-use).",
     "modules/auth_v2/services/mfa_service.py (new), modules/auth_v2/resources/auth_resource.py",
     "🟠 High", 7, "12",
     "User can enroll, login with TOTP, use recovery code, disable MFA. All paths tested.",
     "Medium", "Not Started"),

    (26, "Phase 4 – Advance", "Architecture",
     "API versioning strategy + v3 scaffolding",
     "Decide URL vs header versioning (recommend URL: /api/v2 stable, /api/v3 new). Document deprecation policy "
     "(6-month sunset, deprecation header). Scaffold /api/v3 blueprint structure mirroring v2.",
     "modules/core/api_registry.py, docs/api_versioning.md",
     "🟡 Medium", 4, "14",
     "Versioning policy doc committed. /api/v3/health returns 200.",
     "Low", "Not Started"),

    (27, "Phase 4 – Advance", "Reliability",
     "Webhook retry with exponential backoff + DLQ inspection",
     "Outbound notifier (FCM, email): use Celery autoretry_for=(NetworkError,) with retry_backoff=True, max_retries=5. "
     "Build admin endpoint /admin/dlq to inspect and replay failed tasks.",
     "modules/tasks/, modules/admin/resources/",
     "🟡 Medium", 5, "18",
     "Forced FCM failure retries with backoff; failed-after-max ends in DLQ; replay from admin works.",
     "Medium", "Not Started"),

    (28, "Phase 4 – Advance", "Compliance",
     "Audit log for all money movements",
     "Append-only audit_events table. Service layer hooks: every wallet credit/debit, role change, MFA event, "
     "failed login emits an audit row with actor, action, target, before/after hash, ip, user_agent. Read-only "
     "admin endpoint with filters.",
     "modules/core/audit.py (new), modules/wallet/services/, modules/auth_v2/services/",
     "🟠 High", 7, "24",
     "Every wallet transaction in tests produces a matching audit row. Tampering check (hash chain) verifiable.",
     "Medium", "Not Started"),

    (29, "Phase 4 – Advance", "Documentation",
     "API docs + on-call runbook",
     "Polish OpenAPI specs (Smorest auto-generates; review descriptions, examples). Build docs/runbook.md: "
     "common alerts (webhook lag, DB pool exhausted, Sentry spike), debugging steps, escalation path. Add "
     "ARCHITECTURE.md with current C4 model.",
     "docs/runbook.md, docs/ARCHITECTURE.md, modules/**/resources/",
     "🟡 Medium", 6, "21,28",
     "Runbook covers ≥ 5 alert types. /docs renders cleanly with examples for top 10 endpoints.",
     "Low", "Not Started"),

    (30, "Phase 4 – Advance", "Review",
     "Demo, retrospective, 60-day plan",
     "Live demo: webhook flow, MFA enrollment, fraud rule triggering, reconciliation run. Retro doc: what shipped, "
     "what slipped, what surprised. Draft 60-day plan covering remaining tech-debt + FastAPI feasibility decision "
     "based on Day 20 load-test data.",
     "docs/retro_30d.md, docs/plan_60d.md",
     "🟠 High", 5, "23,24,25,28,29",
     "Demo recorded. Retro and 60-day plan committed. Tag v2.1.0.",
     "Low", "Not Started"),
]

# write headers
for c, h in enumerate(headers, start=1):
    plan.cell(row=1, column=c, value=h)
style_header(plan, 1, len(headers))
plan.freeze_panes = "A2"

# write data
for r, row in enumerate(ROWS, start=2):
    day, phase, category, task, subtasks, files, prio, eff, deps, accept, risk, status = row
    d = working_day(START, day)
    plan.cell(row=r, column=1,  value=day)
    plan.cell(row=r, column=2,  value=d.strftime("%a %Y-%m-%d"))
    plan.cell(row=r, column=3,  value=phase)
    plan.cell(row=r, column=4,  value=category)
    plan.cell(row=r, column=5,  value=task)
    plan.cell(row=r, column=6,  value=subtasks)
    plan.cell(row=r, column=7,  value=files)
    plan.cell(row=r, column=8,  value=prio)
    plan.cell(row=r, column=9,  value=eff)
    plan.cell(row=r, column=10, value=deps)
    plan.cell(row=r, column=11, value=accept)
    plan.cell(row=r, column=12, value=risk)
    plan.cell(row=r, column=13, value=status)

    # phase color on column C
    plan.cell(row=r, column=3).fill = PatternFill("solid", fgColor=PHASE_COLORS[phase])

    for c in range(1, len(headers) + 1):
        cell = plan.cell(row=r, column=c)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = BORDER

    plan.row_dimensions[r].height = 92

# column widths
autosize(plan, [5, 16, 22, 14, 38, 60, 38, 12, 10, 12, 50, 8, 14])

# Status dropdown
status_dv = DataValidation(
    type="list",
    formula1='"Not Started,In Progress,Blocked,In Review,Done"',
    allow_blank=False,
)
status_dv.add(f"M2:M{len(ROWS)+1}")
plan.add_data_validation(status_dv)

# Conditional formatting on Priority column (H)
prio_range = f"H2:H{len(ROWS)+1}"
plan.conditional_formatting.add(
    prio_range,
    FormulaRule(formula=['ISNUMBER(SEARCH("Critical", H2))'],
                fill=PatternFill("solid", fgColor="F8CBAD")),
)
plan.conditional_formatting.add(
    prio_range,
    FormulaRule(formula=['ISNUMBER(SEARCH("High", H2))'],
                fill=PatternFill("solid", fgColor="FFD966")),
)
plan.conditional_formatting.add(
    prio_range,
    FormulaRule(formula=['ISNUMBER(SEARCH("Medium", H2))'],
                fill=PatternFill("solid", fgColor="FFE699")),
)
plan.conditional_formatting.add(
    prio_range,
    FormulaRule(formula=['ISNUMBER(SEARCH("Low", H2))'],
                fill=PatternFill("solid", fgColor="C6EFCE")),
)

# Status column (M)
status_range = f"M2:M{len(ROWS)+1}"
plan.conditional_formatting.add(
    status_range,
    FormulaRule(formula=['M2="Done"'], fill=PatternFill("solid", fgColor="C6EFCE")),
)
plan.conditional_formatting.add(
    status_range,
    FormulaRule(formula=['M2="In Progress"'], fill=PatternFill("solid", fgColor="FFE699")),
)
plan.conditional_formatting.add(
    status_range,
    FormulaRule(formula=['M2="Blocked"'], fill=PatternFill("solid", fgColor="F8CBAD")),
)
plan.conditional_formatting.add(
    status_range,
    FormulaRule(formula=['M2="In Review"'], fill=PatternFill("solid", fgColor="DDEBF7")),
)

# ============================================================
# Sheet 3: Phase Milestones
# ============================================================
ms = wb.create_sheet("Phase Milestones")
ms["A1"] = "Phase Milestones & Deliverables"
ms["A1"].font = TITLE_FONT
ms.merge_cells("A1:E1")

ms_headers = ["Phase", "End-of-Phase Deliverable", "Acceptance Gate", "Owner", "Stakeholder Sign-off"]
for c, h in enumerate(ms_headers, start=1):
    ms.cell(row=3, column=c, value=h)
style_header(ms, 3, len(ms_headers))

ms_rows = [
    ("Phase 1 – Stabilize",
     "v2.0.0-rc1 tag with idempotent webhooks, atomic balance updates, /metrics, Sentry.",
     "Concurrency tests pass 10x; Sentry receives a test exception; /metrics scraped successfully.",
     "Lead Dev",
     "CTO"),
    ("Phase 2 – Harden",
     "v2.0.0-rc2 with green CI, JSON logs, rate limits, security headers, dep scan.",
     "PR cannot merge with failing CI; pip-audit HIGH count = 0; rate limit verified by load test.",
     "Lead Dev",
     "Security"),
    ("Phase 3 – Optimize",
     "Load test baseline doc; query audit closed; async webhooks live; OTel traces visible.",
     "p95 < 500ms at 100 RPS for /wallet endpoints; webhook handler p95 < 100ms.",
     "Lead Dev",
     "CTO + Ops"),
    ("Phase 4 – Advance",
     "v2.1.0: reconciliation, fraud rules v1, MFA, audit log, runbook, 60-day plan.",
     "Demo passes acceptance script; runbook reviewed by ops; audit log immutability verified.",
     "Lead Dev",
     "CTO + Compliance"),
]
for i, row in enumerate(ms_rows, start=4):
    for c, val in enumerate(row, start=1):
        cell = ms.cell(row=i, column=c, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = BORDER
    ms.cell(row=i, column=1).fill = PatternFill("solid", fgColor=PHASE_COLORS[row[0]])
    ms.row_dimensions[i].height = 70

autosize(ms, [22, 55, 55, 16, 22])
ms.freeze_panes = "A4"

# ============================================================
# Sheet 4: Risk Register
# ============================================================
rr = wb.create_sheet("Risk Register")
rr["A1"] = "Risk Register"
rr["A1"].font = TITLE_FONT
rr.merge_cells("A1:G1")

rr_headers = ["ID", "Risk", "Likelihood", "Impact", "Mitigation", "Owner", "Trigger Day"]
for c, h in enumerate(rr_headers, start=1):
    rr.cell(row=3, column=c, value=h)
style_header(rr, 3, len(rr_headers))

risks = [
    ("R1", "Production outage from balance-update refactor (Day 4)",
     "Low", "High",
     "Ship behind feature flag; deploy to staging first; canary 10% of traffic for 24h before full rollout; "
     "have rollback migration ready.",
     "Lead Dev", 4),
    ("R2", "Async webhook switch (Day 18) introduces processing lag visible to users",
     "Medium", "Medium",
     "Keep sync path behind a flag; measure end-to-end latency before flip; budget < 5s from webhook → balance reflected.",
     "Lead Dev", 18),
    ("R3", "Reconciliation surfaces historical discrepancies that look like new bugs",
     "High", "Medium",
     "First run is observe-only (no alerts). Triage backlog separately. Only enable alerting once historical drift is reconciled.",
     "Lead Dev + Ops", 23),
    ("R4", "Fraud rules false-positive blocks legitimate users",
     "Medium", "High",
     "Ship rules in 'shadow mode' (log decisions, do not block) for 5 days. Tune thresholds from real data before enforcing.",
     "Lead Dev", 24),
    ("R5", "MFA rollout breaks existing auth flows",
     "Low", "High",
     "MFA strictly opt-in. Feature flag per-environment. Recovery codes mandatory at enrollment.",
     "Lead Dev", 25),
    ("R6", "Schema migrations on Neon take a lock and stall traffic",
     "Medium", "High",
     "Use CONCURRENTLY for indexes (Day 16). Apply DDL during low-traffic window. Test on staging clone first.",
     "Lead Dev", 16),
    ("R7", "30 days is aggressive — at least one phase will slip",
     "High", "Low",
     "Buffer is implicit in non-merge days. If slipping, drop Phase 4 scope (MFA or fraud) before sacrificing Phase 1/2.",
     "Lead Dev", 14),
    ("R8", "Knowledge gaps from previous owner (undocumented ops runbooks)",
     "High", "Medium",
     "Day 2 includes a money-flow read-through. Day 29 builds the runbook. Schedule 1h with previous owner if reachable.",
     "Lead Dev", 1),
]
for i, row in enumerate(risks, start=4):
    for c, val in enumerate(row, start=1):
        cell = rr.cell(row=i, column=c, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = BORDER
    rr.row_dimensions[i].height = 60
    # color by impact
    impact = row[3]
    color = {"Low": "C6EFCE", "Medium": "FFE699", "High": "F8CBAD"}.get(impact, "FFFFFF")
    rr.cell(row=i, column=4).fill = PatternFill("solid", fgColor=color)

autosize(rr, [6, 50, 12, 10, 60, 18, 12])
rr.freeze_panes = "A4"

# ============================================================
# Sheet 5: KPIs & Success Metrics
# ============================================================
kpi = wb.create_sheet("KPIs & Metrics")
kpi["A1"] = "KPIs & Success Metrics"
kpi["A1"].font = TITLE_FONT
kpi.merge_cells("A1:E1")

kpi_headers = ["Metric", "Baseline (Day 0)", "Target (Day 30)", "Measurement Source", "Notes"]
for c, h in enumerate(kpi_headers, start=1):
    kpi.cell(row=3, column=c, value=h)
style_header(kpi, 3, len(kpi_headers))

kpis = [
    ("Test coverage on modules/wallet", "TBD (capture Day 1)", "≥ 85%", "pytest --cov", "Concurrency suite is the big lift"),
    ("Test coverage on modules/community", "TBD", "≥ 80%", "pytest --cov", ""),
    ("CI pipeline pass rate", "n/a (no CI)", "100% on main, < 10% flake on PR", "GitHub Actions", "Flake budget enforced via @pytest.mark.flaky"),
    ("Webhook handler p95 latency", "TBD", "< 100ms after async switch", "Locust + APM", "Day 18 + Day 20"),
    ("/wallet endpoints p95 latency at 100 RPS", "TBD", "< 500ms", "Locust", "Day 20 baseline"),
    ("Sentry coverage of unhandled exceptions", "0%", "100% in prod-mode env", "Sentry", "Verify with /sentry-debug route"),
    ("Money-flow concurrency tests", "0", "≥ 6 passing", "pytest tests/integration/", "Days 6, 18"),
    ("Number of money-path race conditions", "≥ 2 known", "0 known", "Code review + tests", "Idempotency + balance"),
    ("Reconciliation discrepancies / day", "Unmeasured", "0 unexplained after Day 25", "Daily Celery beat report", "Triage tail of historical drift separately"),
    ("Audit log coverage of money events", "0%", "100%", "Manual sample audit", "Day 28"),
    ("Outstanding HIGH/CRITICAL pip-audit findings", "TBD", "0", "pip-audit in CI", "Day 13"),
    ("Documented on-call runbook", "None", "Covers ≥ 5 alert types", "docs/runbook.md", "Day 29"),
    ("Mean time to detect prod incident", "Unknown", "< 5 min via Sentry/Slack alert", "Sentry alert rules", "After Day 5"),
]
for i, row in enumerate(kpis, start=4):
    for c, val in enumerate(row, start=1):
        cell = kpi.cell(row=i, column=c, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = BORDER
    kpi.row_dimensions[i].height = 36

autosize(kpi, [40, 22, 28, 26, 40])
kpi.freeze_panes = "A4"

# ============================================================
# Sheet 6: Tech Stack Additions
# ============================================================
ts = wb.create_sheet("Tech Stack Additions")
ts["A1"] = "New Tools & Libraries Introduced"
ts["A1"].font = TITLE_FONT
ts.merge_cells("A1:E1")

ts_headers = ["Tool / Library", "Purpose", "Introduced Day", "Add to", "Notes"]
for c, h in enumerate(ts_headers, start=1):
    ts.cell(row=3, column=c, value=h)
style_header(ts, 3, len(ts_headers))

tools = [
    ("sentry-sdk[flask]", "Production error tracking", 5, "requirements.txt + requirements-prod.txt", "Set traces_sample_rate=0.1"),
    ("testcontainers-python", "Real Postgres for integration tests", 6, "requirements.txt (dev)", "Or pytest-postgresql alternative"),
    ("ruff", "Linting + import sorting", 10, "dev tooling", "Replaces flake8 + isort"),
    ("mypy", "Type checking (best-effort)", 10, "dev tooling", "Start with --strict on modules/core only"),
    ("python-json-logger", "Structured JSON logging in prod", 11, "requirements-prod.txt", "Toggle via FLASK_ENV"),
    ("Flask-Limiter", "Rate limiting", 12, "requirements.txt", "Backed by Redis"),
    ("flask-talisman", "Security headers", 13, "requirements.txt", "HSTS, CSP, X-Frame-Options"),
    ("pip-audit", "Dependency vulnerability scan", 13, "CI only", "Block on HIGH"),
    ("trufflehog", "Secret scan in CI", 13, "CI only", "Pre-commit hook also recommended"),
    ("opentelemetry-* suite", "Distributed tracing", 21, "requirements.txt", "Flask + SQLAlchemy + Redis + requests instrumentation"),
    ("locust", "Load testing", 20, "requirements.txt (dev)", "Run from tests/load/"),
    ("pyotp", "TOTP for MFA", 25, "requirements.txt", "Plus qrcode for enrollment QR"),
    ("celery beat", "Scheduled tasks (recon, retries)", 23, "docker-compose service", "Already in requirements"),
]
for i, row in enumerate(tools, start=4):
    for c, val in enumerate(row, start=1):
        cell = ts.cell(row=i, column=c, value=val)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = BORDER
    ts.row_dimensions[i].height = 28

autosize(ts, [28, 36, 16, 36, 40])
ts.freeze_panes = "A4"

# ============================================================
# Sheet 7: Daily Effort Summary (for capacity planning)
# ============================================================
eff = wb.create_sheet("Effort Summary")
eff["A1"] = "Effort Summary (focused hours per day, excludes meetings/review)"
eff["A1"].font = TITLE_FONT
eff.merge_cells("A1:D1")

eff_headers = ["Phase", "Days", "Total Effort (h)", "Avg Effort/Day (h)"]
for c, h in enumerate(eff_headers, start=1):
    eff.cell(row=3, column=c, value=h)
style_header(eff, 3, len(eff_headers))

# compute totals from ROWS
phase_totals = {}
phase_days = {}
for row in ROWS:
    phase = row[1]
    effort = row[7]
    phase_totals[phase] = phase_totals.get(phase, 0) + effort
    phase_days[phase] = phase_days.get(phase, 0) + 1

r = 4
for phase in ["Phase 1 – Stabilize", "Phase 2 – Harden", "Phase 3 – Optimize", "Phase 4 – Advance"]:
    days = phase_days[phase]
    total = phase_totals[phase]
    eff.cell(row=r, column=1, value=phase).fill = PatternFill("solid", fgColor=PHASE_COLORS[phase])
    eff.cell(row=r, column=2, value=days)
    eff.cell(row=r, column=3, value=total)
    eff.cell(row=r, column=4, value=round(total/days, 1))
    for c in range(1, 5):
        eff.cell(row=r, column=c).border = BORDER
    r += 1

eff.cell(row=r, column=1, value="TOTAL").font = Font(bold=True)
eff.cell(row=r, column=2, value=sum(phase_days.values())).font = Font(bold=True)
eff.cell(row=r, column=3, value=sum(phase_totals.values())).font = Font(bold=True)
eff.cell(row=r, column=4, value=round(sum(phase_totals.values()) / sum(phase_days.values()), 1)).font = Font(bold=True)
for c in range(1, 5):
    eff.cell(row=r, column=c).border = BORDER

autosize(eff, [24, 10, 18, 22])

# ---------- save ----------
wb.save(OUT)
print(f"Wrote {OUT}")
